import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import OUTPUT_EXCEL_FILE, OUTPUT_COLUMNS, OUTPUT_FOLDER


class ExcelExporter:
    """Excel导出器类"""

    def __init__(self, filename=None):
        """初始化Excel导出器"""
        self.filename = filename or OUTPUT_EXCEL_FILE
        self.filepath = os.path.join(OUTPUT_FOLDER, self.filename)
        self.workbook = None
        self.worksheet = None
        self._ensure_output_folder()

    def _ensure_output_folder(self):
        """确保输出文件夹存在"""
        if not os.path.exists(OUTPUT_FOLDER):
            os.makedirs(OUTPUT_FOLDER)

    def create_excel(self, data_list):
        """创建并填充Excel文件"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "招投标数据"

        # 设置表头
        self._set_headers()

        # 填充数据
        self._fill_data(data_list)

        # 调整列宽
        self._adjust_column_widths()

        # 保存文件
        self.workbook.save(self.filepath)
        print(f"✓ Excel文件已保存: {self.filepath}")
        return self.filepath

    def _set_headers(self):
        """设置表头（第1行为标题，第2行为列头）"""
        # 第1行：大标题
        title_cell = self.worksheet.cell(row=1, column=1)
        title_cell.value = "全国内分泌配送商联系表"
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        title_alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = title_alignment

        # 合并第1行所有列
        self.worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(OUTPUT_COLUMNS))

        # 第2行：列头
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, column_name in enumerate(OUTPUT_COLUMNS, 1):
            cell = self.worksheet.cell(row=2, column=col_idx)
            cell.value = column_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _fill_data(self, data_list):
        """填充数据到Excel（从第3行开始）"""
        for row_idx, data_dict in enumerate(data_list, 3):  # 从第3行开始
            for col_idx, column_name in enumerate(OUTPUT_COLUMNS, 1):
                cell = self.worksheet.cell(row=row_idx, column=col_idx)
                value = data_dict.get(column_name, "")
                cell.value = value if value else "-"
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _adjust_column_widths(self):
        """自动调整列宽"""
        column_widths = {
            "企业名称": 30,
            "省份": 12,
            "企业经营范围": 40,
            "企业地址": 35,
            "企业法人": 15,
            "企业联系电话": 18,
            "成立日期": 15,
            "营业期限": 18,
            "注册资金": 15,
            "统一社会信用代码": 20,
            "纳税人识别号": 18,
            "实际业务负责人": 18,
            "实际联系号码": 18,
            "代理产品类别": 25,
            "微信/邮箱": 25,
            "配送省份": 30,
            "覆盖地区": 30,
            "覆盖医院": 40
        }

        for col_idx, column_name in enumerate(OUTPUT_COLUMNS, 1):
            column_letter = self.worksheet.cell(row=1, column=col_idx).column_letter
            width = column_widths.get(column_name, 20)
            self.worksheet.column_dimensions[column_letter].width = width

        # 设置行高
        self.worksheet.row_dimensions[1].height = 35  # 标题行
        self.worksheet.row_dimensions[2].height = 30  # 列头行
        for row_idx in range(3, self.worksheet.max_row + 1):  # 数据行
            self.worksheet.row_dimensions[row_idx].height = 25


def export_to_excel(data_list, filename=None):
    """导出数据到Excel的便捷函数"""
    exporter = ExcelExporter(filename)
    return exporter.create_excel(data_list)


if __name__ == "__main__":
    # 测试
    test_data = [
        {
            "企业名称": "测试公司1",
            "省份": "北京",
            "企业经营范围": "医药代理",
            "企业地址": "北京市朝阳区",
            "企业法人": "张三",
            "企业联系电话": "010-12345678"
        }
    ]
    export_to_excel(test_data, "测试.xlsx")
